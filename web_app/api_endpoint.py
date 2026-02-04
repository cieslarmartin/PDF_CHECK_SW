# api_endpoint.py
# API endpointy pro příjem dat z desktop agenta
# Build 41 | © 2025 Ing. Martin Cieślar
# NOVÉ: Licenční systém, device binding, feature flags

from flask import request, jsonify
from database import Database
import logging
import os
import secrets
import time

# Import licenčního systému
try:
    from license_config import (
        LicenseTier, create_license_token, validate_license_token,
        get_tier_features, get_tier_limits, tier_to_string
    )
    from feature_manager import FeatureManager, create_manager_from_api_key
    LICENSE_SYSTEM_AVAILABLE = True
except ImportError:
    LICENSE_SYSTEM_AVAILABLE = False

logger = logging.getLogger(__name__)

# Jednorázové přihlašovací tokeny (agent → web): token -> { 'api_key', 'expires' }
_one_time_login_tokens = {}
_TOKEN_EXPIRY_SEC = 120


def consume_one_time_token(token):
    """
    Pro web: vymění jednorázový token za api_key a licence info. Token se po použití smaže.
    Vrací (api_key, license_info_dict) nebo (None, None) při neplatném/vypršeném tokenu.
    """
    global _one_time_login_tokens
    import time
    token = (token or '').strip()
    if not token:
        return None, None
    now = time.time()
    for t in list(_one_time_login_tokens.keys()):
        if _one_time_login_tokens[t]['expires'] <= now:
            del _one_time_login_tokens[t]
    if token not in _one_time_login_tokens:
        return None, None
    data = _one_time_login_tokens.pop(token)
    if data['expires'] <= now:
        return None, None
    api_key = data['api_key']
    _db = Database()
    license_info = _db.get_user_license(api_key) if api_key else None
    if not license_info:
        return None, None
    return api_key, license_info


def _request_client_info(request):
    """Vrátí (ip_address, machine_id, machine_name) z requestu (headers + remote)."""
    ip = request.headers.get('X-Forwarded-For', request.remote_addr) or ''
    if isinstance(ip, str) and ',' in ip:
        ip = ip.split(',')[0].strip()
    machine_id = (request.headers.get('X-Machine-ID') or '').strip() or None
    machine_name = (request.headers.get('X-Machine-Name') or '').strip() or None
    return ip or None, machine_id, machine_name


def register_api_routes(app):
    """Zaregistruje API endpointy do Flask aplikace"""

    db = Database()

    # =========================================================================
    # BATCH ENDPOINTY (NOVÉ v40)
    # =========================================================================

    @app.route('/api/batch/create', methods=['POST'])
    def create_batch():
        """
        Vytvoří novou dávku pro skupinu souborů

        Headers:
            Authorization: Bearer {API_KEY}

        Body (JSON):
            {
                "batch_name": "Kontrola - 2026-01-29",
                "source_folder": "C:/Documents/Project"
            }

        Returns:
            {
                "success": true,
                "batch_id": "batch_20260129_143256_abc12345"
            }
        """
        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Missing or invalid Authorization header'}), 401

            api_key = auth_header.replace('Bearer ', '').strip()

            if not db.verify_api_key(api_key):
                return jsonify({'error': 'Invalid API key'}), 401

            data = request.get_json() if request.is_json else {}
            batch_name = data.get('batch_name')
            source_folder = data.get('source_folder')

            batch_id = db.create_batch(api_key, batch_name, source_folder)

            if batch_id:
                logger.info(f"Batch vytvořen: {batch_id}")
                return jsonify({
                    'success': True,
                    'batch_id': batch_id
                }), 200
            else:
                return jsonify({'error': 'Failed to create batch'}), 500

        except Exception as e:
            logger.exception(f"Chyba create batch: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/batch/upload', methods=['POST'])
    def upload_batch():
        """
        Nahraje CELÝ batch najednou v jednom requestu.

        Headers:
            Authorization: Bearer {API_KEY}

        Body (JSON):
            {
                "batch_name": "Kontrola - 2026-01-30 12:00",
                "source_folder": "C:/Documents/Project",
                "total_files": 100,
                "results": [
                    {
                        "file_name": "soubor.pdf",
                        "folder": "IO-01/A",
                        "relative_path": "IO-01/A/soubor.pdf",
                        "file_hash": "abc123",
                        "file_size": 12345,
                        "processed_at": "2026-01-30T12:00:00",
                        "results": {...}
                    },
                    ...
                ]
            }
        """
        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Missing or invalid Authorization header'}), 401

            api_key = auth_header.replace('Bearer ', '').strip()

            if not db.verify_api_key(api_key):
                return jsonify({'error': 'Invalid API key'}), 401

            ip_address, machine_id, machine_name = _request_client_info(request)
            if machine_id and db.is_user_device_blocked(api_key, machine_id):
                return jsonify({'error': 'Licence limit reached. Contact support.'}), 403

            if not request.is_json:
                return jsonify({'error': 'Content-Type must be application/json'}), 400

            data = request.get_json()
            batch_name = data.get('batch_name')
            source_folder = data.get('source_folder')
            results = data.get('results', [])

            if not results:
                return jsonify({'error': 'No results provided'}), 400

            total_submitted = len(results)

            # Limit podle licence: použij max_batch_size z licence (nebo tier default)
            license_info = db.get_user_license(api_key)
            tier_name = (license_info or {}).get('tier_name') or ''
            is_trial = str(tier_name).strip().lower() == 'trial'

            # Trial: vázáno na Machine-ID, celkový limit souborů napříč relacemi
            if is_trial:
                if not machine_id or not str(machine_id).strip():
                    return jsonify({
                        'error': 'Zkušební režim vyžaduje identifikaci zařízení (X-Machine-ID). Restartujte agenta.'
                    }), 403
                usage = db.get_trial_usage(machine_id)
                total_so_far = (usage or {}).get('total_files', 0)
                limit = getattr(db, 'TRIAL_LIMIT_TOTAL_FILES', 10)
                if total_so_far >= limit:
                    return jsonify({
                        'error': 'Zkušební limit vyčerpán. Zakupte si prosím licenci.'
                    }), 403
                # Po zpracování batch zvýšíme trial_usage (viz níže)

            max_files = license_info.get('max_batch_size', 5) if license_info else 5
            max_devices = license_info.get('max_devices', 1) if license_info else 1
            if machine_id and not is_trial:
                existing = {d['machine_id'] for d in db.get_user_devices_list(api_key)}
                if machine_id not in existing and max_devices >= 0:
                    if db.count_user_devices_non_blocked(api_key) >= max_devices:
                        return jsonify({'error': 'Licence limit reached. Contact support.'}), 403
                db.upsert_user_device(api_key, machine_id, machine_name)
            # -1 = neomezeno
            if max_files >= 0 and total_submitted > max_files:
                results = results[:max_files]  # slice – zpracuj jen povolený počet

            # Trial: zkontroluj znovu, že po oříznutí nepřekročíme celkový limit
            if is_trial and machine_id:
                usage = db.get_trial_usage(machine_id)
                total_so_far = (usage or {}).get('total_files', 0)
                limit = getattr(db, 'TRIAL_LIMIT_TOTAL_FILES', 10)
                remaining = max(0, limit - total_so_far)
                if remaining <= 0:
                    return jsonify({
                        'error': 'Zkušební limit vyčerpán. Zakupte si prosím licenci.'
                    }), 403
                if len(results) > remaining:
                    results = results[:remaining]

            # Denní kvóta (BASIC 500/den, PRO 1000/den; -1 = neomezeno)
            limits = license_info.get('limits') or {}
            daily_limit = limits.get('daily_files_limit')
            if daily_limit is None:
                tier = LicenseTier(license_info.get('license_tier', 0))
                daily_limit = get_tier_limits(tier).get('daily_files_limit', -1)
            if not is_trial and daily_limit is not None and daily_limit >= 0:
                daily_used = db.get_daily_files_checked(api_key)
                if daily_used + len(results) > daily_limit:
                    return jsonify({
                        'error': 'Denní kvóta vyčerpána. Limit bude obnoven do půlnoci.'
                    }), 403

            # Vytvoř batch
            batch_id = db.create_batch(api_key, batch_name, source_folder)
            if not batch_id:
                return jsonify({'error': 'Failed to create batch'}), 500

            # Ulož výsledky (už oříznuté na max_files)
            saved_count = 0
            for result in results:
                success, _ = db.save_result(api_key, result, batch_id)
                if success:
                    saved_count += 1

            # Trial: zvýš počítadlo pro toto zařízení
            if is_trial and machine_id and saved_count > 0:
                db.increment_trial_usage(machine_id, saved_count)

            # Aktualizuj statistiky batch
            db.update_batch_stats(batch_id)

            total_size_kb = sum((r.get('file_size') or 0) for r in results) // 1024
            db.insert_user_log(
                api_key, 'batch_upload', file_count=saved_count,
                total_size_kb=total_size_kb, ip_address=ip_address, machine_id=machine_id, status='ok'
            )

            is_partial = total_submitted > saved_count

            logger.info(f"Batch upload: {batch_id} - {saved_count}/{total_submitted} souborů (partial={is_partial})")

            resp = {
                'success': True,
                'batch_id': batch_id,
                'saved_count': saved_count,
                'total_count': total_submitted,
                'processed_count': saved_count,
            }
            if is_partial:
                resp['status'] = 'partial'
                resp['message'] = f'Limit {saved_count} souborů překročen. Zpracováno prvních {saved_count}, zbytek ignorován.'

            return jsonify(resp), 200 if not is_partial else 201

        except Exception as e:
            logger.exception(f"Chyba batch upload: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/batch/<batch_id>/finalize', methods=['POST'])
    def finalize_batch(batch_id):
        """
        Finalizuje dávku - aktualizuje statistiky

        Headers:
            Authorization: Bearer {API_KEY}
        """
        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Missing or invalid Authorization header'}), 401

            api_key = auth_header.replace('Bearer ', '').strip()

            if not db.verify_api_key(api_key):
                return jsonify({'error': 'Invalid API key'}), 401

            db.update_batch_stats(batch_id)
            logger.info(f"Batch finalizován: {batch_id}")

            return jsonify({
                'success': True,
                'message': 'Batch finalized'
            }), 200

        except Exception as e:
            logger.exception(f"Chyba finalize batch: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/batch/<batch_id>', methods=['DELETE'])
    def delete_batch(batch_id):
        """
        Smaže dávku – jen pokud patří přihlášenému uživateli (Authorization: Bearer api_key).
        """
        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Přihlášení vyžadováno'}), 401
            api_key = auth_header.replace('Bearer ', '').strip()
            if not db.verify_api_key(api_key):
                return jsonify({'error': 'Neplatný klíč'}), 401

            owner_key = db.get_batch_api_key(batch_id)
            if not owner_key or owner_key != api_key:
                return jsonify({'error': 'Přístup odepřen – dávka nepatří vašemu účtu'}), 403

            if db.delete_batch(batch_id):
                logger.info(f"Batch smazán: {batch_id} (api_key: {api_key[:12]}...)")
                return jsonify({'success': True, 'message': 'Batch deleted'}), 200
            return jsonify({'error': 'Failed to delete batch'}), 500

        except Exception as e:
            logger.exception(f"Chyba delete batch: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/all-data', methods=['DELETE'])
    def delete_all_data():
        """
        Smaže pouze data přihlášeného uživatele (jeho batche a výsledky).
        Vyžaduje Authorization: Bearer api_key. Nikdy nesmaže data jiných uživatelů.
        """
        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Přihlášení vyžadováno'}), 401
            api_key = auth_header.replace('Bearer ', '').strip()
            if not db.verify_api_key(api_key):
                return jsonify({'error': 'Neplatný klíč'}), 401

            deleted = db.delete_all_results_for_api_key(api_key)
            logger.info(f"Smazáno {deleted} záznamů uživatele (api_key: {api_key[:12]}...)")
            return jsonify({'success': True, 'message': f'Deleted {deleted} records'}), 200

        except Exception as e:
            logger.exception(f"Chyba delete all: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    # =========================================================================
    # PŮVODNÍ ENDPOINTY (upravené pro batch)
    # =========================================================================

    @app.route('/api/results', methods=['POST'])
    def receive_results():
        """
        Přijme výsledky z desktop agenta

        Headers:
            Authorization: Bearer {API_KEY}

        Body (JSON):
            {
                "batch_id": "batch_20260129_143256_abc12345",  // NOVÉ - volitelné
                "file_name": "pudorys.pdf",
                "file_path": "subfolder/pudorys.pdf",  // NOVÉ
                "folder": "subfolder",  // NOVÉ
                "file_hash": "a3f8e9...",
                "file_size": 2458632,
                "processed_at": "2025-01-29T12:34:56",
                "results": {...}
            }
        """
        try:
            # Ověř API klíč
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Missing or invalid Authorization header'}), 401

            api_key = auth_header.replace('Bearer ', '').strip()

            if not db.verify_api_key(api_key):
                logger.warning(f"Neplatný API klíč: {api_key[:10]}...")
                return jsonify({'error': 'Invalid API key'}), 401

            # Získej JSON data
            if not request.is_json:
                return jsonify({'error': 'Content-Type must be application/json'}), 400

            data = request.get_json()

            # Validace dat
            if not data.get('file_name'):
                return jsonify({'error': 'Missing file_name'}), 400

            # Získej batch_id (volitelné)
            batch_id = data.get('batch_id')

            # Ulož do databáze
            success, result_id = db.save_result(api_key, data, batch_id)

            if success:
                logger.info(f"Výsledek uložen: {data.get('file_name')} (ID: {result_id}, batch: {batch_id})")
                return jsonify({
                    'success': True,
                    'message': 'Results received and stored',
                    'result_id': result_id
                }), 200
            else:
                logger.error(f"Chyba při ukládání: {result_id}")
                return jsonify({'error': f'Database error: {result_id}'}), 500

        except Exception as e:
            logger.exception(f"Chyba API endpointu: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/auth/verify', methods=['GET'])
    def verify_api_key():
        """Ověří platnost API klíče"""
        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Missing or invalid Authorization header'}), 401

            api_key = auth_header.replace('Bearer ', '').strip()

            if db.verify_api_key(api_key):
                return jsonify({
                    'success': True,
                    'message': 'API key is valid'
                }), 200
            else:
                return jsonify({'error': 'Invalid API key'}), 401

        except Exception as e:
            logger.exception(f"Chyba ověřování: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/auth/user-login', methods=['POST'])
    def user_login():
        """
        Přihlášení uživatele e-mailem a heslem (pro agenta).
        Body (JSON): { "email": "...", "password": "..." }
        Vrátí: { "success": true, "api_key": "...", "user_name": "...", "email": "...", "tier_name": "..." }
        """
        try:
            if not request.is_json:
                return jsonify({'success': False, 'error': 'Content-Type must be application/json'}), 400
            data = request.get_json()
            email = (data.get('email') or '').strip()
            password = data.get('password') or ''
            if not email:
                return jsonify({'success': False, 'error': 'Zadejte e-mail'}), 400
            if not password:
                return jsonify({'success': False, 'error': 'Zadejte heslo'}), 400
            success, result = db.verify_license_password(email, password)
            if not success:
                return jsonify({'success': False, 'error': result}), 401
            api_key = result['api_key']
            ip_address, machine_id, machine_name = _request_client_info(request)
            if machine_id and db.is_user_device_blocked(api_key, machine_id):
                return jsonify({'success': False, 'error': 'Licence limit reached. Contact support.'}), 403
            if machine_id:
                license_info = db.get_user_license(api_key)
                max_devices = (license_info or {}).get('max_devices', 1)
                existing = {d['machine_id'] for d in db.get_user_devices_list(api_key)}
                if machine_id not in existing and max_devices >= 0:
                    if db.count_user_devices_non_blocked(api_key) >= max_devices:
                        return jsonify({'success': False, 'error': 'Licence limit reached. Contact support.'}), 403
                db.upsert_user_device(api_key, machine_id, machine_name)
            db.insert_user_log(api_key, 'login', file_count=0, total_size_kb=0, ip_address=ip_address, machine_id=machine_id, status='ok')
            # max_batch_size z DB (Trial a ostatní tier definice) – agent zobrazí "Trial verze - Limit X souborů"
            return jsonify({
                'success': True,
                'api_key': api_key,
                'user_name': result.get('user_name'),
                'email': result.get('email'),
                'tier': result.get('license_tier', 0),
                'tier_name': result.get('tier_name'),
                'max_batch_size': result.get('max_batch_size'),
            }), 200
        except Exception as e:
            logger.exception(f"Chyba user-login: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500

    @app.route('/api/auth/one-time-login-token', methods=['POST'])
    def one_time_login_token():
        """
        Pro agenta: s platným API klíčem vrátí URL s jednorázovým tokenem.
        Agent otevře tuto URL v prohlížeči – web uživatele automaticky přihlásí.
        Bezpečné: token je jednorázový a platí cca 2 minuty.
        """
        global _one_time_login_tokens
        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'success': False, 'error': 'Přihlášení vyžadováno (Bearer API klíč)'}), 401
            api_key = auth_header.replace('Bearer ', '').strip()
            if not db.verify_api_key(api_key):
                return jsonify({'success': False, 'error': 'Neplatný API klíč'}), 401
            license_info = db.get_user_license(api_key)
            if not license_info or license_info.get('is_expired'):
                return jsonify({'success': False, 'error': 'Licence vypršela nebo není platná'}), 403
            token = secrets.token_urlsafe(32)
            _one_time_login_tokens[token] = {
                'api_key': api_key,
                'expires': time.time() + _TOKEN_EXPIRY_SEC
            }
            # Base URL pro odkaz: na PythonAnywhere vždy HTTPS, jinak z requestu
            host = request.host or request.headers.get('Host', '')
            if 'pythonanywhere.com' in host:
                base_url = f"https://{host}".rstrip('/')
            else:
                base_url = request.url_root.rstrip('/')
            # Přímý vstup do kontroly (/app) s přihlášením – ne na landing
            login_url = f"{base_url}/auth/from-agent-token?login_token={token}"
            return jsonify({'success': True, 'login_url': login_url}), 200
        except Exception as e:
            logger.exception(f"Chyba one-time-login-token: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500

    @app.route('/api/auth/session-from-token', methods=['GET'])
    def session_from_token():
        """
        Výměna jednorázového tokenu (z URL ?login_token=xxx) za údaje uživatele.
        Volá se z webu po otevření odkazu z agenta. Token se po použití zruší.
        """
        global _one_time_login_tokens
        try:
            token = request.args.get('token') or request.args.get('login_token') or ''
            if not token:
                return jsonify({'success': False, 'error': 'Chybí token'}), 400
            now = time.time()
            # Vyčisti vypršené tokeny (mutace, ne přiřazení)
            for t in list(_one_time_login_tokens.keys()):
                if _one_time_login_tokens[t]['expires'] <= now:
                    del _one_time_login_tokens[t]
            if token not in _one_time_login_tokens:
                return jsonify({'success': False, 'error': 'Neplatný nebo vypršený token'}), 401
            data = _one_time_login_tokens.pop(token)
            if data['expires'] <= now:
                return jsonify({'success': False, 'error': 'Token vypršel'}), 401
            api_key = data['api_key']
            license_info = db.get_user_license(api_key)
            if not license_info:
                return jsonify({'success': False, 'error': 'Účet nenalezen'}), 401
            return jsonify({
                'success': True,
                'api_key': api_key,
                'user_name': license_info.get('user_name'),
                'email': license_info.get('email'),
                'tier': license_info.get('license_tier', 0),
                'tier_name': license_info.get('tier_name', 'Free'),
            }), 200
        except Exception as e:
            logger.exception(f"Chyba session-from-token: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500

    @app.route('/api/generate-key', methods=['POST'])
    def generate_key():
        """Vygeneruje nový API klíč (pro testování)"""
        try:
            from database import generate_api_key

            user_name = None
            if request.is_json:
                data = request.get_json()
                user_name = data.get('user_name')

            new_key = generate_api_key()

            if db.create_api_key(new_key, user_name):
                logger.info(f"Nový API klíč vygenerován pro: {user_name or 'Anonymous'}")
                return jsonify({
                    'success': True,
                    'api_key': new_key,
                    'message': 'API key generated successfully'
                }), 200
            else:
                return jsonify({'error': 'Failed to create API key'}), 500

        except Exception as e:
            logger.exception(f"Chyba generování klíče: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/stats', methods=['GET'])
    def get_stats():
        """Vrátí statistiky pro API klíč"""
        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Missing or invalid Authorization header'}), 401

            api_key = auth_header.replace('Bearer ', '').strip()

            if not db.verify_api_key(api_key):
                return jsonify({'error': 'Invalid API key'}), 401

            stats = db.get_statistics(api_key)
            return jsonify(stats), 200

        except Exception as e:
            logger.exception(f"Chyba statistik: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/results/list', methods=['GET'])
    def list_results():
        """Vrátí seznam výsledků pro API klíč"""
        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Missing or invalid Authorization header'}), 401

            api_key = auth_header.replace('Bearer ', '').strip()

            if not db.verify_api_key(api_key):
                return jsonify({'error': 'Invalid API key'}), 401

            limit = int(request.args.get('limit', 100))
            offset = int(request.args.get('offset', 0))

            results = db.get_results_by_api_key(api_key, limit, offset)

            return jsonify({
                'success': True,
                'results': results,
                'count': len(results)
            }), 200

        except Exception as e:
            logger.exception(f"Chyba list results: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    # =========================================================================
    # AGENT RESULTS (pro webové rozhraní) - UPRAVENÉ v40
    # =========================================================================

    @app.route('/api/agent/results', methods=['GET'])
    def get_agent_results():
        """
        Vrátí výsledky JEN pro přihlášeného uživatele (Authorization: Bearer api_key).
        Bez platného přihlášení vrací prázdné batche – žádná data jiných uživatelů.
        Strikní oddělení: tisíce uživatelů, každý vidí jen své kontroly a historii.
        """
        try:
            api_key = None
            auth_header = request.headers.get('Authorization')
            if auth_header and auth_header.startswith('Bearer '):
                api_key = auth_header.replace('Bearer ', '').strip()
                if not db.verify_api_key(api_key):
                    api_key = None

            # Bez platného přihlášení žádná data – striktní oddělení uživatelů
            if not api_key:
                return jsonify({
                    'success': True,
                    'stats': {'total_checks': 0, 'pdf_a3_count': 0, 'batch_count': 0},
                    'license': {'tier': 0, 'tier_name': 'Free'},
                    'batches': []
                }), 200

            batches = db.get_agent_results_grouped(limit=50, api_key=api_key)
            total_files = sum(len(b.get('results', [])) for b in batches)
            pdf_a3_count = sum(
                sum(1 for r in b.get('results', []) if r.get('is_pdf_a3'))
                for b in batches
            )
            lic = db.get_user_license(api_key)
            license_info = {
                'tier': lic.get('license_tier', 0),
                'tier_name': lic.get('tier_name', 'Free')
            } if lic else {'tier': 0, 'tier_name': 'Free'}
            if lic:
                limits = lic.get('limits') or {}
                if limits.get('daily_files_limit') is None:
                    tier = LicenseTier(lic.get('license_tier', 0))
                    limits['daily_files_limit'] = get_tier_limits(tier).get('daily_files_limit', -1)
                license_info['daily_files_used'] = db.get_daily_files_checked(api_key)
                daily_limit = limits.get('daily_files_limit')
                license_info['daily_files_remaining'] = max(0, daily_limit - license_info['daily_files_used']) if daily_limit is not None and daily_limit >= 0 else None
                license_info['daily_files_limit'] = daily_limit

            return jsonify({
                'success': True,
                'stats': {
                    'total_checks': total_files,
                    'pdf_a3_count': pdf_a3_count,
                    'batch_count': len(batches)
                },
                'license': license_info,
                'batches': batches
            }), 200

        except Exception as e:
            logger.exception(f"Chyba agent results: {e}")
            return jsonify({'error': str(e)}), 500

    @app.route('/api/agent/batch/<batch_id>/export', methods=['GET'])
    def export_batch_excel(batch_id):
        """
        Exportuje dávku do Excel (XLSX) – jen vlastní dávku přihlášeného uživatele.
        Vyžaduje Authorization: Bearer api_key. Batch musí patřit tomuto api_key.
        PRO verze: allow_excel_export musí být True (Basic/Trial nemají přístup).
        """
        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Přihlášení vyžadováno'}), 401
            api_key = auth_header.replace('Bearer ', '').strip()
            if not db.verify_api_key(api_key):
                return jsonify({'error': 'Neplatný klíč'}), 401

            lic = db.get_user_license(api_key)
            if not lic or not lic.get('allow_excel_export'):
                return jsonify({'error': 'Export do Excelu je dostupný pouze v PRO verzi.'}), 403

            owner_key = db.get_batch_api_key(batch_id)
            if not owner_key or owner_key != api_key:
                return jsonify({'error': 'Přístup odepřen – dávka nepatří vašemu účtu'}), 403

            results = db.get_batch_results(batch_id)
            if not results:
                return jsonify({'error': 'Batch not found or empty'}), 404

            export_format = request.args.get('format', 'xlsx')

            if export_format == 'json':
                return jsonify({
                    'success': True,
                    'batch_id': batch_id,
                    'results': results
                }), 200

            # Excel export s formátováním
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            import io

            wb = Workbook()
            ws = wb.active
            ws.title = "PDF Check"

            # Styly
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Header
            headers = ['Složka', 'Soubor', 'PDF/A-3', 'Verze', 'Podpis', 'Jméno', 'ČKAIT/ČKA', 'TSA', 'Datum kontroly']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Data
            for row_idx, r in enumerate(results, 2):
                parsed = r.get('parsed_results', {})
                pdf_format = parsed.get('results', {}).get('pdf_format', {})
                signatures = parsed.get('results', {}).get('signatures', [])

                signer = ', '.join(s.get('name', '—') for s in signatures if s.get('name') and s.get('name') != '—') or '—'
                ckait = ', '.join(s.get('ckait_number', '—') for s in signatures if s.get('ckait_number') and s.get('ckait_number') != '—') or '—'
                tsa = 'TSA' if any(s.get('timestamp_valid') for s in signatures) else ('Lokální' if signatures else 'Žádné')

                # Formát data: YYYY-MM-DD HH:MM
                processed_at = r.get('processed_at', '')
                if processed_at:
                    if 'T' in processed_at:
                        processed_at = processed_at.replace('T', ' ')
                    if '.' in processed_at:
                        processed_at = processed_at.split('.')[0]
                    parts = processed_at.split(':')
                    if len(parts) >= 2:
                        processed_at = ':'.join(parts[:2])

                row_data = [
                    r.get('folder_path', '.'),
                    r.get('file_name', ''),
                    'Ano' if r.get('is_pdf_a3') else 'Ne',
                    pdf_format.get('exact_version', ''),
                    'Ano' if signatures else 'Ne',
                    signer,
                    ckait,
                    tsa,
                    processed_at
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border

            # Automatická šířka sloupců podle obsahu
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                # Přidej trochu prostoru
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Zmrazit horní řádek
            ws.freeze_panes = 'A2'

            # Přidat automatický filtr
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"

            # Získej název batch pro filename
            batch_info = db.get_batches()
            batch_name = batch_id
            for b in batch_info:
                if b.get('batch_id') == batch_id:
                    batch_name = b.get('batch_name', batch_id)
                    break

            # Sanitize filename
            safe_name = batch_name.replace(' ', '_').replace('/', '-').replace(':', '-').replace('\\', '-')

            # Uložit do paměti
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            from flask import Response
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename={safe_name}.xlsx'}
            )

        except Exception as e:
            logger.exception(f"Chyba export batch: {e}")
            return jsonify({'error': str(e)}), 500

    @app.route('/api/agent/export-all', methods=['GET'])
    def export_all_excel():
        """
        Exportuje do Excel jen data přihlášeného uživatele.
        Vyžaduje Authorization: Bearer api_key. PRO verze: allow_excel_export musí být True.
        """
        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Přihlášení vyžadováno'}), 401
            api_key = auth_header.replace('Bearer ', '').strip()
            if not db.verify_api_key(api_key):
                return jsonify({'error': 'Neplatný klíč'}), 401

            lic = db.get_user_license(api_key)
            if not lic or not lic.get('allow_excel_export'):
                return jsonify({'error': 'Export do Excelu je dostupný pouze v PRO verzi.'}), 403

            all_batches = db.get_agent_results_grouped(limit=100, api_key=api_key)
            all_results = []
            for batch in all_batches:
                batch_name = batch.get('batch_name', 'Neznámý')
                for r in batch.get('results', []):
                    r['_batch_name'] = batch_name
                    all_results.append(r)

            if not all_results:
                return jsonify({'error': 'No data to export'}), 404

            # Excel export s formátováním
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            import io
            from datetime import datetime

            wb = Workbook()
            ws = wb.active
            ws.title = "PDF Check - Vše"

            # Styly
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Header
            headers = ['Kontrola', 'Složka', 'Soubor', 'PDF/A-3', 'Verze', 'Podpis', 'Jméno', 'ČKAIT/ČKA', 'TSA', 'Datum kontroly']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Data
            for row_idx, r in enumerate(all_results, 2):
                parsed = r.get('parsed_results', {})
                pdf_format = parsed.get('results', {}).get('pdf_format', {})
                signatures = parsed.get('results', {}).get('signatures', [])

                signer = ', '.join(s.get('name', '—') for s in signatures if s.get('name') and s.get('name') != '—') or '—'
                ckait = ', '.join(s.get('ckait_number', '—') for s in signatures if s.get('ckait_number') and s.get('ckait_number') != '—') or '—'
                tsa = 'TSA' if any(s.get('timestamp_valid') for s in signatures) else ('Lokální' if signatures else 'Žádné')

                # Formát data: YYYY-MM-DD HH:MM
                processed_at = r.get('processed_at', '')
                if processed_at:
                    if 'T' in processed_at:
                        processed_at = processed_at.replace('T', ' ')
                    if '.' in processed_at:
                        processed_at = processed_at.split('.')[0]
                    parts = processed_at.split(':')
                    if len(parts) >= 2:
                        processed_at = ':'.join(parts[:2])

                row_data = [
                    r.get('_batch_name', ''),
                    r.get('folder_path', '.'),
                    r.get('file_name', ''),
                    'Ano' if r.get('is_pdf_a3') else 'Ne',
                    pdf_format.get('exact_version', ''),
                    'Ano' if signatures else 'Ne',
                    signer,
                    ckait,
                    tsa,
                    processed_at
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border

            # Automatická šířka sloupců
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Zmrazit horní řádek
            ws.freeze_panes = 'A2'

            # Přidat automatický filtr
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(all_results) + 1}"

            # Filename
            safe_name = f"PDF_Check_Export_{datetime.now().strftime('%Y-%m-%d_%H-%M')}"

            # Uložit do paměti
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            from flask import Response
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename={safe_name}.xlsx'}
            )

        except Exception as e:
            logger.exception(f"Chyba export all: {e}")
            return jsonify({'error': str(e)}), 500

    # =========================================================================
    # LICENČNÍ ENDPOINTY (NOVÉ v41)
    # =========================================================================

    @app.route('/api/auth/validate-license', methods=['POST'])
    def validate_license():
        """
        Validuje licenci a vrátí signed JWT token s permissions

        Body (JSON):
            {
                "api_key": "sk_test_xxx...",
                "hwid": "HWID123...",      // Hardware ID zařízení
                "device_name": "PC-Work",   // Volitelné - název zařízení
                "os_info": "Windows 11"     // Volitelné - info o OS
            }

        Returns:
            {
                "success": true,
                "token": "eyJ...",           // Signed JWT token
                "tier": 2,
                "tier_name": "Pro",
                "features": [...],
                "limits": {...},
                "expires_in": 86400,
                "device_registered": true
            }
        """
        if not LICENSE_SYSTEM_AVAILABLE:
            return jsonify({'error': 'License system not available'}), 503

        try:
            if not request.is_json:
                return jsonify({'error': 'Content-Type must be application/json'}), 400

            data = request.get_json()
            api_key = data.get('api_key')
            hwid = data.get('hwid')
            device_name = data.get('device_name')
            os_info = data.get('os_info')

            if not api_key:
                return jsonify({'error': 'Missing api_key'}), 400

            if not hwid:
                return jsonify({'error': 'Missing hwid'}), 400

            # Validuj zařízení a získej licenci
            valid, result = db.validate_device(api_key, hwid)

            if not valid:
                logger.warning(f"License validation failed for {api_key[:15]}...: {result}")
                return jsonify({
                    'success': False,
                    'error': result,
                    'demo_mode': True
                }), 401

            # Aktualizuj device info pokud jsou poskytnuty
            if device_name or os_info:
                # Update device info v databázi
                pass  # TODO: přidat metodu pro update device info

            # Vytvoř signed token
            tier = LicenseTier(result.get('license_tier', 0))
            token = create_license_token(
                api_key=api_key,
                tier=tier,
                hwid=hwid,
                user_name=result.get('user_name')
            )

            logger.info(f"License validated: {api_key[:15]}... tier={tier_to_string(tier)}")

            return jsonify({
                'success': True,
                'token': token,
                'tier': int(tier),
                'tier_name': tier_to_string(tier),
                'features': result.get('features', []),
                'limits': result.get('limits', {}),
                'expires_in': 24 * 3600,  # Token platí 24 hodin
                'device_registered': True,
                'user_name': result.get('user_name'),
                'days_remaining': result.get('days_remaining', -1)
            }), 200

        except Exception as e:
            logger.exception(f"License validation error: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/auth/register-device', methods=['POST'])
    def register_device():
        """
        Registruje nové zařízení pro API klíč

        Body (JSON):
            {
                "api_key": "sk_test_xxx...",
                "hwid": "HWID123...",
                "device_name": "PC-Work",
                "os_info": "Windows 11"
            }
        """
        if not LICENSE_SYSTEM_AVAILABLE:
            return jsonify({'error': 'License system not available'}), 503

        try:
            if not request.is_json:
                return jsonify({'error': 'Content-Type must be application/json'}), 400

            data = request.get_json()
            api_key = data.get('api_key')
            hwid = data.get('hwid')
            device_name = data.get('device_name')
            os_info = data.get('os_info')

            if not api_key or not hwid:
                return jsonify({'error': 'Missing api_key or hwid'}), 400

            success, message = db.register_device(api_key, hwid, device_name, os_info)

            if success:
                logger.info(f"Device registered: {hwid[:10]}... for {api_key[:15]}...")
                return jsonify({
                    'success': True,
                    'message': message
                }), 200
            else:
                return jsonify({
                    'success': False,
                    'error': message
                }), 400

        except Exception as e:
            logger.exception(f"Device registration error: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/auth/devices', methods=['GET'])
    def list_devices():
        """
        Vrátí seznam aktivních zařízení pro API klíč

        Headers:
            Authorization: Bearer {API_KEY}
        """
        if not LICENSE_SYSTEM_AVAILABLE:
            return jsonify({'error': 'License system not available'}), 503

        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Missing or invalid Authorization header'}), 401

            api_key = auth_header.replace('Bearer ', '').strip()

            if not db.verify_api_key(api_key):
                return jsonify({'error': 'Invalid API key'}), 401

            devices = db.get_active_devices(api_key)
            license_info = db.get_user_license(api_key)

            return jsonify({
                'success': True,
                'devices': devices,
                'max_devices': license_info.get('max_devices', 1) if license_info else 1,
                'active_count': len(devices)
            }), 200

        except Exception as e:
            logger.exception(f"List devices error: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/auth/device/<hwid>', methods=['DELETE'])
    def remove_device(hwid):
        """
        Odstraní zařízení z účtu

        Headers:
            Authorization: Bearer {API_KEY}
        """
        if not LICENSE_SYSTEM_AVAILABLE:
            return jsonify({'error': 'License system not available'}), 503

        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Missing or invalid Authorization header'}), 401

            api_key = auth_header.replace('Bearer ', '').strip()

            if not db.verify_api_key(api_key):
                return jsonify({'error': 'Invalid API key'}), 401

            if db.remove_device(api_key, hwid):
                logger.info(f"Device removed: {hwid[:10]}... from {api_key[:15]}...")
                return jsonify({
                    'success': True,
                    'message': 'Device removed'
                }), 200
            else:
                return jsonify({
                    'success': False,
                    'error': 'Device not found'
                }), 404

        except Exception as e:
            logger.exception(f"Remove device error: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/license/info', methods=['GET'])
    def get_license_info():
        """
        Vrátí kompletní informace o licenci

        Headers:
            Authorization: Bearer {API_KEY}
        """
        if not LICENSE_SYSTEM_AVAILABLE:
            return jsonify({'error': 'License system not available'}), 503

        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Missing or invalid Authorization header'}), 401

            api_key = auth_header.replace('Bearer ', '').strip()

            license_info = db.get_user_license(api_key)

            if not license_info:
                return jsonify({'error': 'Invalid API key'}), 401

            # Když má uživatel tier_id (Trial, Free, Basic… z DB), použij limity z DB – ne hardcoded
            if license_info.get('tier_id'):
                license_info['features'] = []
                license_info['limits'] = {
                    'max_files_per_batch': license_info.get('max_batch_size'),
                    'max_devices': license_info.get('max_devices'),
                }
            else:
                tier = LicenseTier(license_info.get('license_tier', 0))
                license_info['features'] = get_tier_features(tier)
                license_info['limits'] = get_tier_limits(tier)
            if license_info.get('limits', {}).get('max_files_per_batch') is None:
                license_info.setdefault('limits', {})['max_files_per_batch'] = license_info.get('max_batch_size')
            # Denní kvóta – vždy doplnit z get_tier_limits (BASIC 500, PRO 1000)
            limits = license_info.get('limits', {})
            if limits.get('daily_files_limit') is None:
                tier = LicenseTier(license_info.get('license_tier', 0))
                limits['daily_files_limit'] = get_tier_limits(tier).get('daily_files_limit', -1)
            daily_used = db.get_daily_files_checked(api_key)
            license_info['daily_files_used'] = daily_used
            daily_limit = limits.get('daily_files_limit')
            license_info['daily_files_remaining'] = max(0, daily_limit - daily_used) if daily_limit is not None and daily_limit >= 0 else None

            return jsonify({
                'success': True,
                'license': license_info
            }), 200

        except Exception as e:
            logger.exception(f"License info error: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    # =========================================================================
    # RATE LIMITING PRO FREE CHECK (NOVÉ v41)
    # =========================================================================

    @app.route('/api/free-check/status', methods=['GET'])
    def free_check_status():
        """
        Zkontroluje status rate limitu pro free check

        Identifikuje podle IP adresy
        """
        try:
            # Získej IP adresu
            ip = request.headers.get('X-Forwarded-For', request.remote_addr)
            if ',' in ip:
                ip = ip.split(',')[0].strip()

            allowed, remaining, reset_seconds = db.check_rate_limit(
                identifier=ip,
                identifier_type='ip',
                action_type='free_check',
                max_per_hour=3
            )

            return jsonify({
                'success': True,
                'allowed': allowed,
                'remaining': remaining,
                'reset_in_seconds': reset_seconds,
                'limit_per_hour': 3
            }), 200

        except Exception as e:
            logger.exception(f"Free check status error: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/free-check/record', methods=['POST'])
    def record_free_check():
        """
        Zaznamená použití free checku (voláno po úspěšné kontrole)
        """
        try:
            # Získej IP adresu
            ip = request.headers.get('X-Forwarded-For', request.remote_addr)
            if ',' in ip:
                ip = ip.split(',')[0].strip()

            # Nejprve zkontroluj limit
            allowed, remaining, _ = db.check_rate_limit(
                identifier=ip,
                identifier_type='ip',
                action_type='free_check',
                max_per_hour=3
            )

            if not allowed:
                return jsonify({
                    'success': False,
                    'error': 'Rate limit exceeded',
                    'remaining': 0,
                    'upgrade_cta': True
                }), 429

            # Zaznamenej akci
            db.record_rate_limit_action(
                identifier=ip,
                identifier_type='ip',
                action_type='free_check'
            )

            return jsonify({
                'success': True,
                'remaining': remaining - 1
            }), 200

        except Exception as e:
            logger.exception(f"Record free check error: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    # =========================================================================
    # ADMIN ENDPOINTY (pro správu licencí)
    # =========================================================================

    @app.route('/api/admin/create-license', methods=['POST'])
    def admin_create_license():
        """
        Vytvoří nový API klíč s licencí (admin endpoint)

        Body (JSON):
            {
                "admin_key": "admin_secret_key",  // Admin autorizace
                "user_name": "Jan Novák",
                "email": "jan@example.com",
                "tier": 2,                        // 0=Free, 1=Basic, 2=Pro, 3=Enterprise
                "license_days": 365               // Platnost v dnech (null = navždy)
            }
        """
        if not LICENSE_SYSTEM_AVAILABLE:
            return jsonify({'error': 'License system not available'}), 503

        try:
            if not request.is_json:
                return jsonify({'error': 'Content-Type must be application/json'}), 400

            data = request.get_json()

            # Jednoduchá admin autorizace (v produkci nahradit lepším řešením)
            admin_key = data.get('admin_key')
            if admin_key != 'pdfcheck_admin_2025':  # TODO: přesunout do konfigurace
                return jsonify({'error': 'Unauthorized'}), 401

            user_name = data.get('user_name')
            email = data.get('email')
            tier = int(data.get('tier', 0))
            license_days = data.get('license_days')

            # Vygeneruj API klíč
            from database import generate_api_key
            new_key = generate_api_key()

            # Vytvoř záznam s licencí
            if db.create_api_key_with_license(
                api_key=new_key,
                user_name=user_name,
                email=email,
                license_tier=tier,
                license_days=license_days
            ):
                tier_name = tier_to_string(LicenseTier(tier))
                logger.info(f"License created: {new_key[:15]}... tier={tier_name} for {user_name}")

                return jsonify({
                    'success': True,
                    'api_key': new_key,
                    'tier': tier,
                    'tier_name': tier_name,
                    'user_name': user_name,
                    'email': email,
                    'license_days': license_days
                }), 200
            else:
                return jsonify({'error': 'Failed to create license'}), 500

        except Exception as e:
            logger.exception(f"Create license error: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/admin/upgrade-license', methods=['POST'])
    def admin_upgrade_license():
        """
        Upgraduje existující licenci

        Body (JSON):
            {
                "admin_key": "admin_secret_key",
                "api_key": "sk_test_xxx...",
                "new_tier": 2,
                "license_days": 365
            }
        """
        if not LICENSE_SYSTEM_AVAILABLE:
            return jsonify({'error': 'License system not available'}), 503

        try:
            if not request.is_json:
                return jsonify({'error': 'Content-Type must be application/json'}), 400

            data = request.get_json()

            admin_key = data.get('admin_key')
            if admin_key != 'pdfcheck_admin_2025':
                return jsonify({'error': 'Unauthorized'}), 401

            api_key = data.get('api_key')
            new_tier = int(data.get('new_tier', 0))
            license_days = data.get('license_days')

            if not api_key:
                return jsonify({'error': 'Missing api_key'}), 400

            if db.update_license_tier(api_key, new_tier, license_days):
                tier_name = tier_to_string(LicenseTier(new_tier))
                logger.info(f"License upgraded: {api_key[:15]}... to {tier_name}")

                return jsonify({
                    'success': True,
                    'api_key': api_key,
                    'new_tier': new_tier,
                    'tier_name': tier_name
                }), 200
            else:
                return jsonify({'error': 'API key not found'}), 404

        except Exception as e:
            logger.exception(f"Upgrade license error: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/admin/list-licenses', methods=['GET'])
    def admin_list_licenses():
        """
        Vrátí seznam všech licencí (admin endpoint)

        Query params:
            admin_key: admin_secret_key
        """
        if not LICENSE_SYSTEM_AVAILABLE:
            return jsonify({'error': 'License system not available'}), 503

        try:
            admin_key = request.args.get('admin_key')
            if admin_key != 'pdfcheck_admin_2025':
                return jsonify({'error': 'Unauthorized'}), 401

            keys = db.get_all_api_keys()

            # Přidej tier name ke každému klíči
            for key in keys:
                tier = key.get('license_tier', 0)
                key['tier_name'] = tier_to_string(LicenseTier(tier)) if tier else 'Free'

            return jsonify({
                'success': True,
                'licenses': keys,
                'count': len(keys)
            }), 200

        except Exception as e:
            logger.exception(f"List licenses error: {e}")
            return jsonify({'error': 'Internal server error'}), 500

    # =========================================================================
    # DEPLOY (git pull + reload na PythonAnywhere) – volá se jedním klikem z PC
    # =========================================================================
    @app.route('/api/deploy', methods=['GET'])
    def deploy():
        """
        Tajný endpoint: po push z PC sem zavoláte ?token=XXX.
        Na serveru (PA) udělá git pull a zavolá PA API Reload.
        Nastavení na PA: env DEPLOY_TOKEN, PA_USERNAME, PA_API_TOKEN, PA_DOMAIN.
        """
        import subprocess
        token = request.args.get('token', '')
        deploy_token = os.environ.get('DEPLOY_TOKEN', '')
        if not deploy_token or token != deploy_token:
            return jsonify({'ok': False, 'error': 'Unauthorized'}), 403
        repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        out = []
        try:
            r = subprocess.run(
                ['git', 'pull'],
                cwd=repo_root,
                capture_output=True,
                text=True,
                timeout=60,
            )
            out.append(f"git pull: {r.stdout or ''} {r.stderr or ''}".strip() or str(r.returncode))
            if r.returncode != 0:
                return jsonify({'ok': False, 'log': out}), 500
        except Exception as e:
            logger.exception(f"Deploy git pull: {e}")
            return jsonify({'ok': False, 'log': out, 'error': str(e)}), 500
        try:
            username = os.environ.get('PA_USERNAME', '')
            api_token = os.environ.get('PA_API_TOKEN', '')
            domain = os.environ.get('PA_DOMAIN', '')
            if not (username and api_token and domain):
                out.append('Reload skipped: PA_USERNAME/PA_API_TOKEN/PA_DOMAIN not set')
                return jsonify({'ok': True, 'log': out})
            url = f'https://www.pythonanywhere.com/api/v0/user/{username}/webapps/{domain}/reload/'
            import urllib.request
            req = urllib.request.Request(url, data=b'', headers={'Authorization': f'Token {api_token}'}, method='POST')
            with urllib.request.urlopen(req, timeout=30) as resp:
                if resp.status == 200:
                    out.append('Reload OK')
                else:
                    out.append(f'Reload status {resp.status}')
        except Exception as e:
            logger.exception(f"Deploy reload: {e}")
            out.append(f"Reload error: {e}")
        return jsonify({'ok': True, 'log': out}), 200

    logger.info("API endpointy zaregistrovány (v41 s licenčním systémem)")
